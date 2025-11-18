import csv
import datetime
import pandas as pd
from openpyxl.utils import get_column_letter
import os
import sys


def ruta_absoluta_archivo(nombre_archivo):
    # Si está compilado con PyInstaller
    if getattr(sys, "frozen", False):
        # Primero intenta en la carpeta del ejecutable
        ruta_base = os.path.dirname(sys.executable)
        ruta_archivo = os.path.join(ruta_base, nombre_archivo)
        if os.path.exists(ruta_archivo):
            return ruta_archivo
        # Si no existe, intenta en _MEIPASS (por si acaso)
        if hasattr(sys, "_MEIPASS"):
            ruta_archivo = os.path.join(sys._MEIPASS, nombre_archivo)
            if os.path.exists(ruta_archivo):
                return ruta_archivo
        # Si no existe en ninguna, retorna la ruta en la carpeta del ejecutable
        return os.path.join(ruta_base, nombre_archivo)
    else:
        # Ejecutando como script Python
        ruta_base = os.path.dirname(__file__)
        return os.path.join(ruta_base, nombre_archivo)


archivo_entrada = ruta_absoluta_archivo("asistencia.csv")
archivo_reporte_diario = ruta_absoluta_archivo("reporte_diario.xlsx")
archivo_reporte_estadistico = ruta_absoluta_archivo("reporte_estadistico.csv")
archivo_intervalo = ruta_absoluta_archivo("intervalo.csv")
horarios = {
    "Elizabeth 1": ("08:00", "18:00"),
    "Orlando 3": ("08:00", "18:00"),
    "Principe 2": ("07:00", "19:00"),  # No confirmado
    "Juan 10": ("07:00", "19:00"),  # No confirmado
    "Chino 4": ("07:00", "19:00"),  # No confirmado
    "Vallejo 7": ("07:00", "19:00"),  # No confirmado
    "Mendez 6": ("07:00", "19:00"),  # No confirmado
    "Teofilo 11": ("07:00", "19:00"),  # No confirmado
    "Nicole 13": ("08:00", "18:00"),  # No confirmado
}


# Funcion - retorna tuple - con trabajadores extraidos del csv
def obtener_trabajadores():
    set_trabajadores = set()

    with open(archivo_entrada, newline="", encoding="utf-8") as archivo:
        lector = csv.DictReader(archivo)
        for fila in lector:
            set_trabajadores.add(fila["Nombre"])

    return tuple(set_trabajadores)


# Funcion - retorna tuple - con trabajadores excluyendo algunos
def obtener_trabajadores_con_exclusion():
    lista_trabajadores = obtener_trabajadores()
    nombres_excluidos = {"Frank 8", "Miguel 9"}

    trabajadores_filtrados = [
        nombre for nombre in lista_trabajadores if nombre not in nombres_excluidos
    ]
    return tuple(trabajadores_filtrados)


# Funcion - retorna objeto - con trabajadores como claves y listas vacías como valores
def crear_obj_trabajadores(tupla):
    return {item: [] for item in tupla}


# Funcion - retorna lista - con días laborables desde el intervalo definido en el CSV
def dias_trabajables_desde_intervalo_csv():
    with open(archivo_intervalo, mode="r", encoding="utf-8") as archivo:
        lector = csv.reader(archivo)
        next(lector)
        fila = next(lector)
        fecha_inicio = datetime.datetime.strptime(fila[0], "%d/%m/%Y").date()

    fecha_fin = datetime.date.today() - datetime.timedelta(days=1)

    dias_laborables = []
    fecha_actual = fecha_inicio

    while fecha_actual <= fecha_fin:
        if fecha_actual.weekday() < 5:  # lunes a sábado
            dias_laborables.append(fecha_actual.strftime("%d/%m/%Y"))
        fecha_actual += datetime.timedelta(days=1)

    return dias_laborables


# Funcion - retorna objeto - con registros de asistencia por trabajador
def agregar_registros_a_obj(obj):
    with open(archivo_entrada, newline="", encoding="utf-8") as archivo:
        lector = csv.DictReader(archivo)
        for fila in lector:
            nombre = fila["Nombre"]
            if nombre in obj:
                registro_filtrado = {
                    "Tiempo": fila["Tiempo"],
                    "Estado": fila["Estado"],
                }
                obj[nombre].append(registro_filtrado)
    return obj


# Funcion - retorna objeto - con reporte de asistencia por trabajador y día
def generar_reporte_asistencia(obj_asistencia, dias_trabajables):
    reporte = {}

    for trabajador, registros in obj_asistencia.items():
        reporte[trabajador] = []

        for dia in dias_trabajables:
            entrada = None
            salida = None

            for reg in registros:
                try:
                    fecha_hora = datetime.datetime.strptime(
                        reg["Tiempo"], "%d/%m/%Y %H:%M:%S"
                    )
                except ValueError:
                    continue  # saltar registros mal formateados

                if fecha_hora.strftime("%d/%m/%Y") == dia:
                    if reg["Estado"] == "Entrada" and (
                        not entrada or fecha_hora < entrada
                    ):
                        entrada = fecha_hora
                    elif reg["Estado"] == "Salida" and (
                        not salida or fecha_hora > salida
                    ):
                        salida = fecha_hora

            observaciones = ""
            tiempo_total = ""
            horas_perdidas = ""
            horas_extras = ""

            if entrada and salida:
                diff = salida - entrada
                horas, rem = divmod(diff.seconds, 3600)
                minutos = rem // 60
                tiempo_total = f"{horas}h {minutos}m"

                # Calcular horas perdidas y extras
                if trabajador in horarios:
                    h_entrada, h_salida = horarios[trabajador]
                    hora_entrada_teorica = datetime.datetime.strptime(
                        f"{dia} {h_entrada}", "%d/%m/%Y %H:%M"
                    )
                    hora_salida_teorica = datetime.datetime.strptime(
                        f"{dia} {h_salida}", "%d/%m/%Y %H:%M"
                    )

                    # Horas perdidas (por tardanza)
                    if entrada > hora_entrada_teorica:
                        retraso = entrada - hora_entrada_teorica
                        rh, rm = divmod(retraso.seconds, 3600)
                        minutos = rm // 60
                        horas_perdidas = f"{rh}h {minutos}m"

                    # Horas extras (por salir más tarde)
                    if salida > hora_salida_teorica:
                        extra = salida - hora_salida_teorica
                        eh, em = divmod(extra.seconds, 3600)
                        minutos = em // 60
                        horas_extras = f"{eh}h {minutos}m"

            elif not entrada and not salida:
                observaciones = "Sin marcas"
            elif not entrada:
                observaciones = "Faltó entrada"
            elif not salida:
                observaciones = "Faltó salida"

            reporte[trabajador].append(
                {
                    "Fecha": dia,
                    "Hora Entrada": entrada.strftime("%H:%M") if entrada else "",
                    "Hora Salida": salida.strftime("%H:%M") if salida else "",
                    "Tiempo Total": tiempo_total,
                    "Horas Extras": horas_extras,
                    "Horas Perdidas": horas_perdidas,
                    "Obs.": observaciones,
                }
            )

    return reporte


# Función - retorna objeto - con resumen estadístico
def generar_reporte_estadisticas(reporte):
    estadisticas = []

    for trabajador, dias in reporte.items():
        inasistencias = 0
        tardanzas = 0
        dias_cumplidos = 0
        total_horas_extra = datetime.timedelta()
        total_horas_perdidas = datetime.timedelta()

        for dia in dias:
            tiempo_total = dia["Tiempo Total"]
            hora_perdida = dia["Horas Perdidas"]
            hora_extra = dia["Horas Extras"]

            if not tiempo_total:
                inasistencias += 1
                continue

            if hora_perdida and hora_perdida != "0h 0m":
                tardanzas += 1
            else:
                # Si no hay horas perdidas y hay registro, se cuenta como día cumplido
                dias_cumplidos += 1

            if hora_perdida and hora_perdida != "0h 0m":
                h, m = map(int, hora_perdida.replace("h", "").replace("m", "").split())
                total_horas_perdidas += datetime.timedelta(hours=h, minutes=m)

            if hora_extra and hora_extra != "0h 0m":
                h, m = map(int, hora_extra.replace("h", "").replace("m", "").split())
                total_horas_extra += datetime.timedelta(hours=h, minutes=m)

        # Calcular diferencia con signo
        diferencia = total_horas_extra - total_horas_perdidas
        segundos = int(diferencia.total_seconds())
        signo = "+" if segundos > 0 else "-" if segundos < 0 else ""
        segundos_abs = abs(segundos)
        horas, minutos = divmod(segundos_abs // 60, 60)
        diferencia_formateada = f"{signo}{horas}h {minutos}m"

        def formatear(td):
            total_seg = int(td.total_seconds())
            h, m = divmod(total_seg // 60, 60)
            return f"{h}h {m}m"

        estadisticas.append(
            [
                trabajador,
                inasistencias,
                tardanzas,
                dias_cumplidos,
                formatear(total_horas_extra),
                formatear(total_horas_perdidas),
                diferencia_formateada,
            ]
        )

    return estadisticas


# Funcion - imprime - los registros de agregar_registros_a_obj()
def mostrar_registros_por_trabajador(obj):
    for trabajador, registros in obj.items():
        print(f"\n🧑 Trabajador: {trabajador}")
        if not registros:
            print("  (Sin registros)")
        else:
            for fila in registros:
                tiempo = fila["Tiempo"]
                estado = fila["Estado"]
                print(f"  • {tiempo} | {estado}")


# Funcion - imprime - el reporte de generar_reporte_asistencia()
def mostrar_asistencia_formateado(reporte):
    for trabajador, dias in reporte.items():
        print(f"\n🧑 {trabajador}")
        for dia in dias:
            print(
                f"📅 {dia['Fecha']}: "
                f"Entrada: {dia['Hora Entrada']} - "
                f"Salida: {dia['Hora Salida']} - "
                f"Total: {dia['Tiempo Total']} - "
                f"Perdidas: {dia['Horas Perdidas']} - "
                f"Extras: {dia['Horas Extras']} - "
                f"Obs.: {dia['Obs.']}"
            )


# Funcion - imprime - el reporte de generar_reporte_estadisticas()
def mostrar_estadisticas_formateadas(estadisticas):
    headers = [
        "👤 Trabajador",
        "❌ Inasistencias",
        "⏰ Tardanzas",
        "🕒 Horas Extra",
        "⌛ Horas Perdidas",
        "📊 Diferencia",
    ]

    col_widths = [
        max(len(str(row[i])) for row in estadisticas + [headers]) + 2
        for i in range(len(headers))
    ]

    header_row = "".join(
        header.ljust(col_widths[i]) for i, header in enumerate(headers)
    )
    print("\n" + header_row)
    print("-" * sum(col_widths))

    for fila in estadisticas:
        row_str = "".join(str(fila[i]).ljust(col_widths[i]) for i in range(len(fila)))
        print(row_str)

    print()


# Funcion - genera excel - reporte diario
def generar_excel_reporte_diario(reporte_final, estadisticas_final):
    with pd.ExcelWriter(archivo_reporte_diario, engine="openpyxl") as writer:
        # Crear DataFrame de las estadísticas
        columnas = [
            "Trabajador",
            "Inasistencias",
            "Tardanzas",
            "Días Cumplidos",
            "Horas Extra",
            "Horas Perdidas",
            "Diferencia",
        ]
        df_estadisticas = pd.DataFrame(estadisticas_final, columns=columnas)
        df_estadisticas.to_excel(writer, sheet_name="Resumen", index=False)

        # Escribir los registros de cada trabajador
        for trabajador, registros in reporte_final.items():
            df = pd.DataFrame(registros).fillna("-")
            df.to_excel(writer, sheet_name=trabajador[:31], index=False)

        # Ajustar el ancho de columnas de todas las hojas
        for sheet_name in writer.book.sheetnames:
            worksheet = writer.book[sheet_name]
            for col_idx, col in enumerate(
                worksheet.iter_cols(1, worksheet.max_column), 1
            ):
                max_length = max(
                    (len(str(cell.value)) for cell in col if cell.value is not None),
                    default=0,
                )
                adjusted_width = max_length + 2
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = adjusted_width


# Funcion - main
def main():
    trabajadores = obtener_trabajadores_con_exclusion()
    dias_laborables = dias_trabajables_desde_intervalo_csv()
    obj_trabajadores = crear_obj_trabajadores(trabajadores)
    objt_trabajadores_asistencia = agregar_registros_a_obj(obj_trabajadores)
    reporte_final = generar_reporte_asistencia(
        objt_trabajadores_asistencia, dias_laborables
    )
    estadisticas_final = generar_reporte_estadisticas(reporte_final)

    print("\nGenerando reporte diario...")
    generar_excel_reporte_diario(reporte_final, estadisticas_final)
    print(f"Reporte diario generado: {archivo_reporte_diario}")


if __name__ == "__main__":
    main()
